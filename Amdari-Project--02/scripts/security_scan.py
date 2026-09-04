#!/usr/bin/env python3
"""
security_scan.py

Run Gitleaks, Semgrep, and Bandit against a repository, save each tool's
native JSON output, merge findings into a severity-ranked Excel report,
and generate a deduplicated unified CSV vulnerability inventory.

Usage:
    python security_scan.py
    python security_scan.py --repo /path/to/repo --output-dir security_reports

Requirements:
    - Python 3.9+
    - gitleaks available on PATH
    - semgrep available on PATH
    - bandit available on PATH
    - openpyxl installed via pip (pip install openpyxl)

The script intentionally does not install third-party scanning tools.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import shutil
import subprocess   # nosec B404
import sys
from collections import Counter
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Iterable

SEVERITY_RANK = {
    "CRITICAL": 4,
    "HIGH": 3,
    "MEDIUM": 2,
    "LOW": 1,
    "INFO": 0,
    "UNKNOWN": -1,
}

TOOL_ORDER = ("gitleaks", "semgrep", "bandit")


@dataclass
class Finding:
    tool: str
    severity: str
    confidence: str
    rule_id: str
    title: str
    message: str
    path: str
    line: Any
    column: Any
    cwe: str
    source: str

    @property
    def severity_rank(self) -> int:
        return SEVERITY_RANK.get(self.severity.upper(), -1)


def run_command(
    cmd: list[str],
    *,
    cwd: Path,
    output_file: Path | None = None,
) -> tuple[int, str]:
    """Run a scanner and optionally save stdout to a file."""
    proc = subprocess.run(   # nosec B603
        cmd,
        cwd=str(cwd),
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        check=False,
    )
    stdout = proc.stdout or ""
    if output_file is not None:
        output_file.write_text(stdout, encoding="utf-8")
    return proc.returncode, stdout


def ensure_tool(name: str) -> str:
    """Return the executable path or stop with a useful error."""
    path = shutil.which(name)
    if not path:
        raise RuntimeError(
            f"Required tool '{name}' was not found on PATH. "
            f"Install it and rerun the scan."
        )
    return path


def parse_json_file(path: Path) -> Any:
    if not path.exists() or not path.read_text(encoding="utf-8").strip():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def normalize_severity(value: Any, default: str = "UNKNOWN") -> str:
    text = str(value or "").strip().upper()
    if text in {"CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"}:
        return text
    if text in {"ERROR", "BLOCKER"}:
        return "CRITICAL"
    if text in {"WARNING", "WARN"}:
        return "MEDIUM"
    return default


def normalize_confidence(value: Any) -> str:
    text = str(value or "").strip().upper()
    return text if text else "UNKNOWN"


def rel_path(repo: Path, path_value: Any) -> str:
    raw = str(path_value or "")
    if not raw:
        return ""
    candidate = Path(raw)
    if candidate.is_absolute():
        try:
            return str(candidate.relative_to(repo))
        except ValueError:
            return str(candidate)
    return raw


def scan_gitleaks(
    repo: Path,
    output_dir: Path,
) -> list[Finding]:
    """
    Gitleaks JSON is stored in gitleaks.json.
    Exit code 1 means leaks were found and is expected.
    """
    output_json = output_dir / "gitleaks.json"
    tool = ensure_tool("gitleaks")

    cmd = [
        tool,
        "detect",
        "--source", str(repo),
        "--report-format", "json",
        "--report-path", str(output_json),
        "--no-banner",
    ]
    rc, combined = run_command(cmd, cwd=repo)

    if not output_json.exists():
        output_json.write_text(
            json.dumps({
                "tool": "gitleaks",
                "returncode": rc,
                "output": combined,
            }, indent=2),
            encoding="utf-8",
        )

    raw = parse_json_file(output_json)
    if not isinstance(raw, list):
        return []

    findings: list[Finding] = []
    for item in raw:
        path = rel_path(
            repo,
            item.get("File") or item.get("file"),
        )
        line = item.get("StartLine") or item.get("startLine")
        column = item.get("StartColumn") or item.get("startColumn")

        findings.append(
            Finding(
                tool="gitleaks",
                severity="CRITICAL",
                confidence="HIGH",
                rule_id=str(
                    item.get("RuleID")
                    or item.get("ruleID")
                    or "gitleaks-secret"
                ),
                title=str(
                    item.get("Description")
                    or "Potential secret detected"
                ),
                message=str(
                    item.get("Description")
                    or "Potential secret detected"
                ),
                path=path,
                line=line,
                column=column,
                cwe="",
                source="gitleaks",
            )
        )
    return findings


def scan_semgrep(
    repo: Path,
    output_dir: Path,
) -> list[Finding]:
    """
    Semgrep JSON is stored in semgrep.json.
    Non-zero status is accepted because scanner behavior can vary by ruleset.
    """
    output_json = output_dir / "semgrep.json"
    tool = ensure_tool("semgrep")

    cmd = [
        tool,
        "--config", "auto",
        "--json",
        "--output", str(output_json),
        str(repo),
    ]
    rc, combined = run_command(cmd, cwd=repo)

    if not output_json.exists():
        output_json.write_text(
            json.dumps({
                "results": [],
                "errors": [{
                    "message": combined,
                    "returncode": rc,
                }],
            }, indent=2),
            encoding="utf-8",
        )

    raw = parse_json_file(output_json)
    results = raw.get("results", []) if isinstance(raw, dict) else []

    findings: list[Finding] = []
    for item in results:
        extra = item.get("extra") or {}
        metadata = extra.get("metadata") or {}
        path = rel_path(
            repo,
            item.get("path"),
        )

        severity = normalize_severity(
            extra.get("severity")
            or metadata.get("severity")
            or "MEDIUM"
        )

        confidence = normalize_confidence(
            extra.get("confidence")
            or metadata.get("confidence")
            or "MEDIUM"
        )

        cwe_value = metadata.get("cwe") or metadata.get("cwe_id") or ""
        if isinstance(cwe_value, list):
            cwe_value = ", ".join(str(v) for v in cwe_value)

        message = (
            extra.get("message")
            or extra.get("metavars")
            or item.get("check_id")
            or "Semgrep finding"
        )

        if not isinstance(message, str):
            message = json.dumps(message, sort_keys=True)

        findings.append(
            Finding(
                tool="semgrep",
                severity=severity,
                confidence=confidence,
                rule_id=str(
                    item.get("check_id")
                    or "semgrep"
                ),
                title=str(
                    metadata.get("shortDescription")
                    or item.get("check_id")
                    or "Semgrep finding"
                ),
                message=message,
                path=path,
                line=(item.get("start") or {}).get("line"),
                column=(item.get("start") or {}).get("col"),
                cwe=str(cwe_value),
                source="semgrep",
            )
        )
    return findings


def scan_bandit(
    repo: Path,
    output_dir: Path,
) -> list[Finding]:
    """
    Bandit JSON is stored in bandit.json.
    """
    output_json = output_dir / "bandit.json"
    tool = ensure_tool("bandit")

    cmd = [
        tool,
        "-r", str(repo),
        "-f", "json",
        "-o", str(output_json),
    ]
    rc, combined = run_command(cmd, cwd=repo)

    if not output_json.exists():
        output_json.write_text(
            json.dumps({
                "results": [],
                "errors": [{
                    "message": combined,
                    "returncode": rc,
                }],
            }, indent=2),
            encoding="utf-8",
        )

    raw = parse_json_file(output_json)
    results = raw.get("results", []) if isinstance(raw, dict) else []

    findings: list[Finding] = []
    for item in results:
        findings.append(
            Finding(
                tool="bandit",
                severity=normalize_severity(
                    item.get("issue_severity"),
                    default="MEDIUM",
                ),
                confidence=normalize_confidence(
                    item.get("issue_confidence")
                ),
                rule_id=str(
                    item.get("test_id")
                    or "bandit"
                ),
                title=str(
                    item.get("test_name")
                    or "Bandit finding"
                ),
                message=str(
                    item.get("issue_text")
                    or "Bandit finding"
                ),
                path=rel_path(
                    repo,
                    item.get("filename"),
                ),
                line=item.get("line_number"),
                column=item.get("col_offset"),
                cwe=str(
                    (item.get("issue_cwe") or {}).get("id", "")
                ),
                source="bandit",
            )
        )
    return findings


def merge_findings(
    findings: Iterable[Finding],
) -> list[Finding]:
    """Rank findings Critical -> Low, then tool/path/line for determinism."""
    return sorted(
        list(findings),
        key=lambda item: (
            -item.severity_rank,
            TOOL_ORDER.index(item.tool)
            if item.tool in TOOL_ORDER
            else 999,
            item.path.lower(),
            int(item.line or 0),
            item.rule_id.lower(),
        ),
    )


def write_aggregated_csv(
    findings: Iterable[Finding],
    output_path: Path,
) -> None:
    """
    Deduplicate and aggregate findings by file and line, then save to a CSV
    compatible with the aggregate.py format.
    """
    severity_map = {
        "CRITICAL": 4,
        "HIGH": 3,
        "ERROR": 3,
        "MEDIUM": 2,
        "WARNING": 2,
        "LOW": 1,
        "INFO": 1,
        "UNKNOWN": 0,
    }
    reverse_map = {
        4: "Critical",
        3: "High",
        2: "Medium",
        1: "Low",
        0: "Info",
    }

    vulnerabilities: dict[str, dict[str, Any]] = {}

    for item in findings:
        file_path = item.path or "unknown"
        line_val = item.line if item.line is not None else 0
        key = f"{file_path}:{line_val}"
        num_sev = severity_map.get(str(item.severity).upper(), 2)
        tool_name = item.tool.capitalize() if item.tool else "Unknown"
        rule_name = item.rule_id or "unknown"
        desc = (item.message or item.title or "").strip().replace("\n", " ")

        if key in vulnerabilities:
            existing = vulnerabilities[key]
            existing["severity_num"] = max(existing["severity_num"], num_sev)
            if tool_name not in existing["tools"]:
                existing["tools"] += f", {tool_name}"
                existing["rules"] += f" | {rule_name}"
        else:
            vulnerabilities[key] = {
                "file": file_path,
                "line": line_val,
                "tools": tool_name,
                "rules": rule_name,
                "severity_num": num_sev,
                "description": desc,
            }

    sorted_vulns = sorted(
        vulnerabilities.values(),
        key=lambda x: x["severity_num"],
        reverse=True,
    )

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Severity",
            "File",
            "Line",
            "Tools",
            "Rules",
            "Description",
            "Triage Status",
        ])
        for v in sorted_vulns:
            writer.writerow([
                reverse_map.get(v["severity_num"], "Low"),
                v["file"],
                v["line"],
                v["tools"],
                v["rules"],
                v["description"],
                "Unreviewed",
            ])


def build_workbook(
    findings: list[Finding],
    repo: Path,
    output_path: Path,
) -> None:
    """
    Build an XLSX workbook using openpyxl.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(
            "ERROR: 'openpyxl' is required to build the workbook. "
            "Install via: pip install openpyxl",
            file=sys.stderr,
        )
        sys.exit(1)

    wb = Workbook()

    # --- 1. Summary Sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"

    fill_title = PatternFill(
        start_color="FF1F2937",
        end_color="FF1F2937",
        fill_type="solid",
    )
    fill_header = PatternFill(
        start_color="FF111827",
        end_color="FF111827",
        fill_type="solid",
    )
    fill_light_grey = PatternFill(
        start_color="FFE5E7EB",
        end_color="FFE5E7EB",
        fill_type="solid",
    )
    fill_critical = PatternFill(
        start_color="FFFECACA",
        end_color="FFFECACA",
        fill_type="solid",
    )
    fill_high = PatternFill(
        start_color="FFFED7AA",
        end_color="FFFED7AA",
        fill_type="solid",
    )
    fill_medium = PatternFill(
        start_color="FFFEF3C7",
        end_color="FFFEF3C7",
        fill_type="solid",
    )
    fill_low = PatternFill(
        start_color="FFDCFCE7",
        end_color="FFDCFCE7",
        fill_type="solid",
    )

    font_white_bold = Font(color="FFFFFFFF", bold=True)
    font_bold = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_wrap = Alignment(wrap_text=True, vertical="top")

    # Title Banner
    ws_summary.merge_cells("A1:F1")
    title_cell = ws_summary["A1"]
    title_cell.value = "Security Scan Report"
    title_cell.fill = fill_title
    title_cell.font = Font(color="FFFFFFFF", bold=True, size=16)
    title_cell.alignment = align_center

    total = len(findings)
    counts = Counter(f.severity for f in findings)

    # Repository & Stats Table
    summary_data = [
        ("Repository", str(repo.resolve())),
        ("Total findings", total),
        ("Critical", counts.get("CRITICAL", 0)),
        ("High", counts.get("HIGH", 0)),
        ("Medium", counts.get("MEDIUM", 0)),
        ("Low", counts.get("LOW", 0)),
        ("Generated by", "security_scan.py"),
    ]

    for i, (k, v) in enumerate(summary_data, start=3):
        label_cell = ws_summary.cell(row=i, column=1, value=k)
        label_cell.font = font_bold
        label_cell.fill = fill_light_grey
        ws_summary.cell(row=i, column=2, value=v)

    # Tool Breakdown Table
    tool_data = [
        ("Tool", "Findings"),
        ("Gitleaks", sum(f.tool == "gitleaks" for f in findings)),
        ("Semgrep", sum(f.tool == "semgrep" for f in findings)),
        ("Bandit", sum(f.tool == "bandit" for f in findings)),
    ]

    for i, (k, v) in enumerate(tool_data, start=3):
        col1 = ws_summary.cell(row=i, column=4, value=k)
        col2 = ws_summary.cell(row=i, column=5, value=v)
        if i == 3:
            col1.fill = fill_header
            col1.font = font_white_bold
            col2.fill = fill_header
            col2.font = font_white_bold

    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 55
    ws_summary.column_dimensions["D"].width = 18
    ws_summary.column_dimensions["E"].width = 18
    ws_summary.freeze_panes = "A3"

    # --- 2. Findings Sheet ---
    ws_report = wb.create_sheet("Findings")

    headers = [
        "Rank", "Severity", "Confidence", "Tool", "Rule ID", "Title",
        "Message", "File", "Line", "Column", "CWE", "Source"
    ]
    ws_report.append(headers)

    for col_num in range(1, len(headers) + 1):
        header_cell = ws_report.cell(row=1, column=col_num)
        header_cell.fill = fill_header
        header_cell.font = font_white_bold
        header_cell.alignment = align_center

    for index, finding in enumerate(findings, start=1):
        row_data = [
            index,
            finding.severity,
            finding.confidence,
            finding.tool,
            finding.rule_id,
            finding.title,
            finding.message,
            finding.path,
            finding.line,
            finding.column,
            finding.cwe,
            finding.source,
        ]
        ws_report.append(row_data)
        current_row = ws_report.max_row

        row_fill = None
        if finding.severity == "CRITICAL":
            row_fill = fill_critical
        elif finding.severity == "HIGH":
            row_fill = fill_high
        elif finding.severity == "MEDIUM":
            row_fill = fill_medium
        elif finding.severity == "LOW":
            row_fill = fill_low

        for col_num in range(1, len(headers) + 1):
            cell = ws_report.cell(row=current_row, column=col_num)
            cell.alignment = align_wrap
            if row_fill:
                cell.fill = row_fill
            if finding.severity in ("CRITICAL", "HIGH") and col_num == 2:
                cell.font = font_bold

    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        if col_letter in ["F", "G", "H"]:
            ws_report.column_dimensions[col_letter].width = 34
        else:
            ws_report.column_dimensions[col_letter].width = 15

    ws_report.freeze_panes = "A2"
    wb.save(str(output_path))


def write_merged_json(
    findings: list[Finding],
    output_path: Path,
) -> None:
    output_path.write_text(
        json.dumps(
            {
                "repository": str(Path(".").resolve()),
                "finding_count": len(findings),
                "findings": [
                    {
                        **asdict(finding),
                        "severity_rank": finding.severity_rank,
                    }
                    for finding in findings
                ],
            },
            indent=2,
            default=str,
        ),
        encoding="utf-8",
    )


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Run Gitleaks, Semgrep and Bandit, save JSON output, "
            "generate a severity-ranked XLSX report, and export an "
            "aggregated CSV inventory."
        )
    )
    parser.add_argument(
        "--repo",
        default=".",
        help="Repository to scan (default: current directory).",
    )
    parser.add_argument(
        "--output-dir",
        default="security_reports",
        help="Directory for JSON, XLSX, and CSV outputs.",
    )
    args = parser.parse_args()

    repo = Path(args.repo).resolve()
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    if not repo.exists():
        print(
            f"ERROR: repository path does not exist: {repo}",
            file=sys.stderr,
        )
        return 2

    all_findings: list[Finding] = []

    scanners = [
        ("gitleaks", scan_gitleaks),
        ("semgrep", scan_semgrep),
        ("bandit", scan_bandit),
    ]

    failures: list[str] = []

    for name, scanner in scanners:
        try:
            all_findings.extend(
                scanner(repo, output_dir)
            )
        except Exception as exc:
            failures.append(
                f"{name}: {exc}"
            )
            print(
                f"[ERROR] {name}: {exc}",
                file=sys.stderr,
            )

    ranked = merge_findings(
        all_findings
    )

    # 1. Machine-readable merged JSON
    write_merged_json(
        ranked,
        output_dir / "merged_findings.json",
    )

    # 2. Aggregated & Deduplicated CSV (compatible with aggregate.py)
    csv_path = output_dir / "unified_inventory.csv"
    write_aggregated_csv(
        ranked,
        csv_path,
    )

    # 3. Formatted openpyxl XLSX workbook
    workbook_path = (
        output_dir / "security_scan_report.xlsx"
    )
    build_workbook(
        ranked,
        repo,
        workbook_path,
    )

    print(
        f"Scanned repository: {repo}"
    )
    print(
        f"JSON outputs: {output_dir}"
    )
    print(
        f"Aggregated CSV inventory: {csv_path}"
    )
    print(
        f"Spreadsheet report: {workbook_path}"
    )

    if failures:
        print(
            "\nScanner/tool failures:",
            file=sys.stderr,
        )
        for failure in failures:
            print(
                f"  - {failure}",
                file=sys.stderr,
            )
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())